import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


def _header(ws, df_src, hdr_color="1F3864"):
    for ci, cn in enumerate(df_src.columns, 1):
        c = ws.cell(row=1, column=ci, value=cn)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=hdr_color)
        c.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"), bottom=Side(style="thin"))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _row_fill(status):
    if "✅" in str(status):
        return "E2EFDA"
    if "⚠" in str(status):
        return "FFF2CC"
    if "❌" in str(status):
        return "FFE0E0"
    return "F2F7FF"


def salvar_relatorio_prevoo(registros, output_dir=None):
    import pandas as pd
    if output_dir is None:
        output_dir = os.getcwd()
    df = pd.DataFrame(registros)
    wb = Workbook()
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws1 = wb.active
    ws1.title = "Dados Completos"
    _header(ws1, df)
    for ri, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        fill = PatternFill("solid", fgColor=_row_fill(row[2] if len(row) > 2 else ""))
        for ci, val in enumerate(row, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.border = thin
            c.fill = fill
            c.alignment = center if ci <= 12 else left_al

    widths = [8, 30, 22, 14, 14, 16, 8, 14, 14, 16, 16, 55, 16, 20, 14, 14, 12, 36, 18, 30, 40, 40]
    for i, w in enumerate(widths[:len(df.columns)], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    cols_res = ["Operadora", "PDF", "Nome (JSON)", "Código do Cliente", "Contrato (JSON)",
                "Referência (mês/ano)", "Data Emissão", "Vencimento", "Nota Fiscal (Nº)",
                "NFF", "Valor Total (R$)", "Status"]
    cols_res = [c for c in cols_res if c in df.columns]
    df_res = df[cols_res].copy() if not df.empty else pd.DataFrame(columns=cols_res)

    ws2 = wb.create_sheet("Resumo Financeiro")
    _header(ws2, df_res, "2E4057")
    total_valor = 0.0
    for ri, row in enumerate(dataframe_to_rows(df_res, index=False, header=False), 2):
        fill = PatternFill("solid", fgColor=_row_fill(row[-1] if row else ""))
        for ci, val in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.border = thin
            c.fill = fill
            c.alignment = center
        try:
            vi = cols_res.index("Valor Total (R$)")
            total_valor += float(str(row[vi]).replace(".", "").replace(",", "."))
        except Exception:
            pass

    tr = len(df_res) + 2
    if "Valor Total (R$)" in cols_res:
        vi = cols_res.index("Valor Total (R$)") + 1
        ws2.cell(row=tr, column=vi-1, value="TOTAL GERAL").font = Font(bold=True)
        tc = ws2.cell(row=tr, column=vi,
                      value=f"R$ {total_valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        tc.font = Font(bold=True, color="1F3864", size=12)
        tc.fill = PatternFill("solid", fgColor="BDD7EE")
        tc.border = thin
        tc.alignment = center

    for i in range(1, len(cols_res) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 24

    ws3 = wb.create_sheet("Problemas")
    df_prob = df[df["Status"].str.contains("❌|⚠", na=False)] if not df.empty else pd.DataFrame()
    if not df_prob.empty:
        _header(ws3, df_prob, "C0392B")
        for ri, row in enumerate(dataframe_to_rows(df_prob, index=False, header=False), 2):
            for ci, val in enumerate(row, 1):
                c = ws3.cell(row=ri, column=ci, value=val)
                c.border = thin
                c.fill = PatternFill("solid", fgColor="FDEDEC")
                c.alignment = center
        for i in range(1, len(df_prob.columns) + 1):
            ws3.column_dimensions[get_column_letter(i)].width = 22
    else:
        ws3.cell(row=1, column=1, value="✅ Todos os PDFs processados sem problemas!").font = Font(bold=True, color="27AE60", size=12)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    caminho = os.path.join(output_dir, f"relatorio_unificado_{ts}.xlsx")
    wb.save(caminho)
    return caminho


def gerar_relatorio_excel(resultados, output_dir=None):
    if output_dir is None:
        output_dir = os.getcwd()
    if not resultados:
        return None
    import pandas as pd
    df = pd.DataFrame(resultados)
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Processamento"
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal="center", vertical="center")

    for cn, col_name in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=cn, value=col_name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4F81BD")
        c.border = thin
        c.alignment = center

    for ri, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = thin
            c.alignment = center

    for col in ws.columns:
        mx = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = mx + 2

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    caminho = os.path.join(output_dir, f"relatorio_macro_{ts}.xlsx")
    wb.save(caminho)
    return caminho
